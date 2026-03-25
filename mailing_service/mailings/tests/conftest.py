from pathlib import Path

import openpyxl
import pytest


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    """Valid XLSX file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["external_id", "user_id", "email", "subject", "message"])
    for i in range(1, 4):
        ws.append(
            [f"ext-{i}", f"user-{i}", f"test{i}@example.com", f"Subject {i}", f"Message {i}"]
        )
    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def xlsx_with_errors(tmp_path: Path) -> Path:
    """XLSX file with 2 valid and 2 invalid rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["external_id", "user_id", "email", "subject", "message"])
    ws.append(["ext-1", "user-1", "test1@example.com", "Subject 1", "Message 1"])
    ws.append(["ext-2", "user-2", "not-an-email", "Subject 2", "Message 2"])
    ws.append(["ext-3", "", "test3@example.com", "Subject 3", "Message 3"])
    ws.append(["ext-4", "user-4", "test4@example.com", "Subject 4", "Message 4"])
    path = tmp_path / "errors.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def empty_xlsx(tmp_path: Path) -> Path:
    """XLSX file with headers only, no data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["external_id", "user_id", "email", "subject", "message"])
    path = tmp_path / "empty.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def xlsx_bad_headers(tmp_path: Path) -> Path:
    """XLSX file with incorrect column headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "user", "mail", "subj", "msg"])
    ws.append(["1", "user-1", "test@example.com", "Subj", "Msg"])
    path = tmp_path / "bad_headers.xlsx"
    wb.save(path)
    return path
