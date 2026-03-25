from __future__ import annotations

import dataclasses
from pathlib import Path

import structlog
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from openpyxl import load_workbook

from mailings.models import MailingRecord
from mailings.services.email import send_email

logger = structlog.get_logger(__name__)

EXPECTED_HEADERS = ["external_id", "user_id", "email", "subject", "message"]


@dataclasses.dataclass
class ImportStats:
    """Statistics of the import process."""

    total_rows: int = 0
    created: int = 0
    skipped: int = 0
    errors: int = 0
    error_details: list[str] = dataclasses.field(default_factory=list)


def _clean_cell(value: object) -> str:
    """openpyxl returns numeric cells as float; normalize to str."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _validate_row(data: dict[str, str], row_num: int) -> str | None:
    """
    Validate a row of data.

    Parameters
    ----------
    data: dict[str, str]
        Data to validate.
    row_num: int
        Number of the row.

    Returns
    -------
    str | None
        Error message if the row is invalid, None if the row is valid.
    """
    for field in EXPECTED_HEADERS:
        if not data.get(field):
            return f"Row {row_num}: missing required field '{field}'"
    try:
        validate_email(data["email"])
    except ValidationError:
        return f"Row {row_num}: invalid email '{data['email']}'"
    return None


def _save_batch(rows: list[dict[str, str]], stats: ImportStats, *, dry_run: bool) -> None:
    """Filter duplicates, bulk-create new records, and dispatch email tasks."""
    existing = set(
        MailingRecord.objects.filter(
            external_id__in=[r["external_id"] for r in rows],
        ).values_list("external_id", flat=True)
    )

    new_rows = [r for r in rows if r["external_id"] not in existing]
    stats.skipped += len(rows) - len(new_rows)

    if not new_rows:
        return

    if dry_run:
        stats.created += len(new_rows)
        return

    MailingRecord.objects.bulk_create(
        [MailingRecord(**r) for r in new_rows],
        ignore_conflicts=True,
    )

    created_qs = MailingRecord.objects.filter(
        external_id__in=[r["external_id"] for r in new_rows],
        status=MailingRecord.Status.PENDING,
    )
    stats.created += created_qs.count()

    for pk in created_qs.values_list("pk", flat=True):
        send_email.delay(pk)


def import_from_xlsx(
    file_path: Path,
    *,
    batch_size: int = 500,
    dry_run: bool = False,
) -> ImportStats:
    """Import mailing records from an XLSX file and dispatch email tasks."""
    stats = ImportStats()
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)

    if raw_headers is None:
        wb.close()
        raise ValueError("File is empty or has no header row")

    headers = [_clean_cell(h).lower() for h in raw_headers]
    if headers != EXPECTED_HEADERS:
        wb.close()
        raise ValueError(f"Invalid headers: expected {EXPECTED_HEADERS}, got {headers}")

    batch: list[dict[str, str]] = []

    for row_num, row in enumerate(rows_iter, start=2):
        stats.total_rows += 1
        data = dict(zip(headers, (_clean_cell(v) for v in row)))

        error = _validate_row(data, row_num)
        if error:
            stats.errors += 1
            stats.error_details.append(error)
            continue

        batch.append(data)
        if len(batch) >= batch_size:
            _save_batch(batch, stats, dry_run=dry_run)
            batch.clear()

    if batch:
        _save_batch(batch, stats, dry_run=dry_run)

    wb.close()
    return stats
